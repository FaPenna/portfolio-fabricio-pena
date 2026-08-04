# -*- coding: utf-8 -*-
"""
Global Controle — Sistema web integrado para gerenciamento de demandas e
controle de ponto eletrônico (REP-P, Portaria nº 671/2021).
TCC — Fabricio Carneiro Pena — Faculdade Projeção / Globalweb Corp.

Execução:  python app.py  →  http://localhost:5000
Usuários seed (senha: 123456):
  admin@globalweb.com.br      (Administrador)
  mariana.souza@globalweb.com.br (Gestora)
  fabricio.pena@globalweb.com.br (Colaborador)
"""
import hashlib
import io
import os
import sqlite3
from datetime import datetime, timedelta, date
from functools import wraps

from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, send_file, g)
from werkzeug.security import generate_password_hash, check_password_hash

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'global_controle.db')

app = Flask(__name__)
app.secret_key = 'global-controle-tcc-2026'

# RN006 — fluxo sequencial de status das demandas
PROXIMO_STATUS = {'pendente': 'andamento', 'andamento': 'concluida'}

SEQUENCIA_PONTO = ['entrada', 'pausa', 'retorno', 'saida']
ROTULO_PONTO = {'entrada': 'Entrada', 'pausa': 'Pausa',
                'retorno': 'Retorno', 'saida': 'Saída'}

MSG = {
    'MSG001': 'Registro de ponto efetuado com sucesso.',
    'MSG002': 'Demanda salva com sucesso.',
    'MSG003': 'Usuário salvo com sucesso.',
    'MSG004': 'Não foi possível gerar o arquivo. Tente novamente.',
    'MSG005': 'Relatório gerado com sucesso.',
    'MSG006': 'Sequência de marcação inválida. Verifique o último registro.',
    'MSG007': 'E-mail ou senha inválidos.',
}


# ---------------------------------------------------------------- banco
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute('PRAGMA foreign_keys = ON')
    return g.db


@app.teardown_appcontext
def close_db(_=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()


SCHEMA = """
CREATE TABLE IF NOT EXISTS DBUsuario (
    id_usuario    INTEGER PRIMARY KEY AUTOINCREMENT,
    nome          VARCHAR(150) NOT NULL,
    email         VARCHAR(150) NOT NULL UNIQUE,
    cpf           VARCHAR(14)  NOT NULL,
    senha         VARCHAR(255) NOT NULL,
    perfil        VARCHAR(20)  NOT NULL CHECK (perfil IN ('colaborador','gestor','admin')),
    ativo         BOOLEAN      NOT NULL DEFAULT 1,
    data_cadastro DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS DBDemanda (
    id_demanda     INTEGER PRIMARY KEY AUTOINCREMENT,
    titulo         VARCHAR(200) NOT NULL,
    descricao      TEXT,
    prioridade     VARCHAR(10)  NOT NULL CHECK (prioridade IN ('baixa','media','alta')),
    status         VARCHAR(20)  NOT NULL DEFAULT 'pendente'
                   CHECK (status IN ('pendente','andamento','concluida')),
    prazo          DATE,
    id_responsavel INTEGER NOT NULL REFERENCES DBUsuario(id_usuario),
    id_criador     INTEGER NOT NULL REFERENCES DBUsuario(id_usuario),
    data_criacao   DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
    data_conclusao DATETIME
);
CREATE TABLE IF NOT EXISTS DBRegistroPonto (
    id_registro INTEGER PRIMARY KEY AUTOINCREMENT,
    id_usuario  INTEGER NOT NULL REFERENCES DBUsuario(id_usuario),
    tipo        VARCHAR(10) NOT NULL CHECK (tipo IN ('entrada','pausa','retorno','saida')),
    data_hora   DATETIME NOT NULL,
    nsr         INTEGER NOT NULL,
    hash        VARCHAR(255) NOT NULL
);
CREATE TABLE IF NOT EXISTS DBRelatorio (
    id_relatorio   INTEGER PRIMARY KEY AUTOINCREMENT,
    tipo           VARCHAR(20) NOT NULL CHECK (tipo IN ('ponto','demandas')),
    periodo_inicio DATE NOT NULL,
    periodo_fim    DATE NOT NULL,
    id_gerador     INTEGER NOT NULL REFERENCES DBUsuario(id_usuario),
    data_geracao   DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
    arquivo        VARCHAR(255)
);
"""


def gerar_hash_registro(id_usuario, tipo, data_hora, nsr):
    """RN003 — hash de integridade que garante a imutabilidade da marcação."""
    base = f'{id_usuario}|{tipo}|{data_hora}|{nsr}|GLOBALCONTROLE'
    return hashlib.sha256(base.encode('utf-8')).hexdigest()


def proximo_nsr(db, id_usuario):
    row = db.execute('SELECT MAX(nsr) AS m FROM DBRegistroPonto WHERE id_usuario=?',
                     (id_usuario,)).fetchone()
    return (row['m'] or 0) + 1


def inserir_ponto(db, id_usuario, tipo, dt):
    nsr = proximo_nsr(db, id_usuario)
    h = gerar_hash_registro(id_usuario, tipo, dt.strftime('%Y-%m-%d %H:%M:%S'), nsr)
    db.execute('INSERT INTO DBRegistroPonto (id_usuario, tipo, data_hora, nsr, hash) '
               'VALUES (?,?,?,?,?)',
               (id_usuario, tipo, dt.strftime('%Y-%m-%d %H:%M:%S'), nsr, h))


# ---------------------------------------------------------------- seed
def migrar_esquema(db):
    """Atualiza bancos criados em versões anteriores do sistema."""
    colunas = [r['name'] for r in db.execute('PRAGMA table_info(DBDemanda)')]
    if 'id_gestor' in colunas and 'id_criador' not in colunas:
        db.execute('ALTER TABLE DBDemanda RENAME COLUMN id_gestor TO id_criador')
        db.commit()
        print('Migração aplicada: DBDemanda.id_gestor -> id_criador')


def seed():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.executescript(SCHEMA)
    migrar_esquema(db)
    if db.execute('SELECT COUNT(*) c FROM DBUsuario').fetchone()['c'] > 0:
        db.close()
        return

    senha = generate_password_hash('123456')
    usuarios = [
        # id 1-3: administração e gestão
        ('Carlos Andrade', 'admin@globalweb.com.br', '111.444.777-35', senha, 'admin'),
        ('Mariana Souza', 'mariana.souza@globalweb.com.br', '222.555.888-46', senha, 'gestor'),
        ('Ricardo Tavares', 'ricardo.tavares@globalweb.com.br', '777.000.333-91', senha, 'gestor'),
        # id 4-9: colaboradores
        ('Fabricio Carneiro Pena', 'fabricio.pena@globalweb.com.br', '333.666.999-57', senha, 'colaborador'),
        ('Juliana Ribeiro', 'juliana.ribeiro@globalweb.com.br', '444.777.000-68', senha, 'colaborador'),
        ('Pedro Henrique Lima', 'pedro.lima@globalweb.com.br', '555.888.111-79', senha, 'colaborador'),
        ('Amanda Castro', 'amanda.castro@globalweb.com.br', '666.999.222-80', senha, 'colaborador'),
        ('Lucas Almeida', 'lucas.almeida@globalweb.com.br', '888.111.444-02', senha, 'colaborador'),
        ('Beatriz Nogueira', 'beatriz.nogueira@globalweb.com.br', '999.222.555-13', senha, 'colaborador'),
    ]
    db.executemany('INSERT INTO DBUsuario (nome,email,cpf,senha,perfil) VALUES (?,?,?,?,?)',
                   usuarios)

    # --- registros de ponto: últimos 30 dias úteis para os 6 colaboradores
    hoje = datetime.now().replace(second=0, microsecond=0)
    colaboradores = [2, 3, 4, 5, 6, 7, 8, 9]   # gestores também registram ponto (UC004)
    d = hoje - timedelta(days=30)
    while d.date() < hoje.date():
        if d.weekday() < 5:
            for uid in colaboradores:
                if (uid + d.day) % 17 == 0:        # falta ocasional p/ dar realismo
                    continue
                desloc = (uid * 7 + d.day) % 23    # pequenas variações por pessoa/dia
                base = d.replace(hour=8, minute=desloc)
                inserir_ponto(db, uid, 'entrada', base)
                inserir_ponto(db, uid, 'pausa', base.replace(hour=12, minute=(desloc + 5) % 59))
                inserir_ponto(db, uid, 'retorno', base.replace(hour=13, minute=(desloc + 11) % 59))
                inserir_ponto(db, uid, 'saida', base.replace(hour=17, minute=(desloc + 34) % 59))
        d += timedelta(days=1)
    # hoje: jornadas em estágios diferentes (bom para os prints de Dashboard e Equipe)
    if hoje.weekday() < 5:
        inserir_ponto(db, 4, 'entrada', hoje.replace(hour=8, minute=2))
        inserir_ponto(db, 4, 'pausa', hoje.replace(hour=12, minute=8))
        inserir_ponto(db, 4, 'retorno', hoje.replace(hour=13, minute=4))
        inserir_ponto(db, 5, 'entrada', hoje.replace(hour=8, minute=17))
        inserir_ponto(db, 6, 'entrada', hoje.replace(hour=8, minute=41))
        inserir_ponto(db, 6, 'pausa', hoje.replace(hour=12, minute=2))
        inserir_ponto(db, 7, 'entrada', hoje.replace(hour=8, minute=55))
        inserir_ponto(db, 8, 'entrada', hoje.replace(hour=7, minute=58))
        inserir_ponto(db, 8, 'pausa', hoje.replace(hour=12, minute=14))
        inserir_ponto(db, 8, 'retorno', hoje.replace(hour=13, minute=9))
        inserir_ponto(db, 2, 'entrada', hoje.replace(hour=8, minute=6))
        inserir_ponto(db, 2, 'pausa', hoje.replace(hour=12, minute=11))
        inserir_ponto(db, 2, 'retorno', hoje.replace(hour=13, minute=2))
        inserir_ponto(db, 3, 'entrada', hoje.replace(hour=8, minute=24))
        # Beatriz (9) sem registro hoje — aparece como "Sem registro" na equipe

    # --- demandas: distribuídas entre os 6 colaboradores e os 2 gestores
    pz = lambda n: (hoje + timedelta(days=n)).strftime('%Y-%m-%d')
    cc = lambda n: (hoje - timedelta(days=n)).strftime('%Y-%m-%d %H:%M:%S')
    demandas = [
        # (titulo, descricao, prioridade, status, prazo, responsavel, criador, concluida_ha_dias)
        ('Atualizar dashboard BI Poupatempo Lote 5', 'Aplicar identidade visual Globalweb e incluir filtro de ausências.', 'alta', 'andamento', pz(2), 4, 2, None),
        ('Documentar rotina de exportação AFD', 'Descrever o leiaute do arquivo conforme Portaria nº 671/2021.', 'media', 'andamento', pz(4), 4, 2, None),
        ('Auditar registros de ponto de maio', 'Validar NSR e hash dos registros do período.', 'media', 'concluida', pz(-3), 4, 2, 3),
        ('Homologar módulo de relatórios', 'Testar exportação Excel com dados reais do mês.', 'baixa', 'concluida', pz(-8), 4, 3, 8),
        ('Revisar chamados 4biz do grupo HR', 'Conferir SLA dos chamados abertos na última semana.', 'alta', 'pendente', pz(1), 5, 2, None),
        ('Testar fluxo de recuperação de senha', 'Executar roteiro de testes funcionais do módulo de autenticação.', 'baixa', 'concluida', pz(-1), 5, 2, 1),
        ('Mapear integrações do portal interno', 'Levantar APIs consumidas pelo portal do colaborador.', 'media', 'andamento', pz(6), 5, 3, None),
        ('Elaborar manual do colaborador remoto', 'Guia rápido de uso do Global Controle para novos colaboradores.', 'baixa', 'pendente', pz(7), 6, 2, None),
        ('Consolidar indicadores semanais da equipe', 'Gerar planilha de horas trabalhadas x demandas concluídas.', 'alta', 'pendente', pz(-1), 6, 3, None),
        ('Migrar scripts de carga para Python 3.12', 'Atualizar dependências e validar execução agendada.', 'media', 'concluida', pz(-5), 6, 3, 5),
        ('Atualizar base de conhecimento Aggrega', 'Migrar artigos para o novo portal interno.', 'media', 'andamento', pz(5), 7, 2, None),
        ('Revisar política de acessos do SGBD', 'Conferir perfis e privilégios no banco de homologação.', 'alta', 'andamento', pz(3), 7, 3, None),
        ('Padronizar templates de e-mail corporativo', 'Aplicar nova identidade nos comunicados internos.', 'baixa', 'concluida', pz(-10), 7, 2, 10),
        ('Inventariar ativos de TI do polo Brasília', 'Atualizar planilha de notebooks e periféricos.', 'media', 'pendente', pz(8), 8, 3, None),
        ('Apoiar onboarding dos novos estagiários', 'Preparar ambiente e acessos da turma de junho.', 'alta', 'andamento', pz(2), 8, 2, None),
        ('Validar backup incremental do servidor', 'Executar restauração de teste do último snapshot.', 'alta', 'concluida', pz(-2), 8, 3, 2),
        ('Catalogar licenças de software vencendo', 'Listar renovações do terceiro trimestre.', 'media', 'pendente', pz(10), 9, 2, None),
        ('Revisar documentação da API de ponto', 'Atualizar exemplos de requisição e respostas.', 'baixa', 'andamento', pz(9), 9, 3, None),
        ('Conferir folha de espelhos de abril', 'Comparar espelhos assinados com o AFD exportado.', 'media', 'concluida', pz(-15), 9, 2, 15),
    ]
    for t, desc, pri, st, prazo, resp, gest, ha in demandas:
        db.execute('INSERT INTO DBDemanda (titulo,descricao,prioridade,status,prazo,'
                   'id_responsavel,id_criador,data_conclusao) VALUES (?,?,?,?,?,?,?,?)',
                   (t, desc, pri, st, prazo, resp, gest, cc(ha) if ha is not None else None))

    # --- histórico de relatórios
    rel = [('ponto', 30, 1, 2), ('demandas', 15, 1, 3), ('ponto', 7, 0, 2), ('demandas', 30, 2, 1)]
    for tipo, dini, dfim, ger in rel:
        db.execute('INSERT INTO DBRelatorio (tipo,periodo_inicio,periodo_fim,id_gerador,arquivo) '
                   'VALUES (?,?,?,?,?)',
                   (tipo, (hoje - timedelta(days=dini)).strftime('%Y-%m-%d'),
                    (hoje - timedelta(days=dfim)).strftime('%Y-%m-%d'),
                    ger, f'relatorio_{tipo}.xlsx'))

    db.commit()
    db.close()


# ---------------------------------------------------------------- auth
def login_obrigatorio(perfis=None):
    def deco(f):
        @wraps(f)
        def wrapper(*a, **kw):
            if 'uid' not in session:
                return redirect(url_for('login'))
            if perfis and session.get('perfil') not in perfis:
                flash('Acesso restrito ao seu perfil.', 'warning')
                return redirect(url_for('dashboard'))
            return f(*a, **kw)
        return wrapper
    return deco


@app.context_processor
def inject_globais():
    DIAS = ['segunda-feira', 'terça-feira', 'quarta-feira', 'quinta-feira',
            'sexta-feira', 'sábado', 'domingo']
    MESES = ['janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho', 'julho',
             'agosto', 'setembro', 'outubro', 'novembro', 'dezembro']
    agora = datetime.now()
    data_pt = f'{DIAS[agora.weekday()]}, {agora.day:02d} de {MESES[agora.month - 1]} de {agora.year}'
    return {'agora': agora, 'data_extenso': data_pt, 'ROTULO_PONTO': ROTULO_PONTO}


# ---------------------------------------------------------------- UC001 login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        db = get_db()
        u = db.execute('SELECT * FROM DBUsuario WHERE email=? AND ativo=1',
                       (request.form['email'].strip().lower(),)).fetchone()
        if u and check_password_hash(u['senha'], request.form['senha']):
            session.update(uid=u['id_usuario'], nome=u['nome'], perfil=u['perfil'])
            return redirect(url_for('dashboard'))
        flash(MSG['MSG007'], 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ---------------------------------------------------------------- helpers ponto
def marcacoes_do_dia(db, uid, dia=None):
    dia = dia or date.today()
    return db.execute(
        "SELECT * FROM DBRegistroPonto WHERE id_usuario=? AND date(data_hora)=? "
        'ORDER BY nsr', (uid, dia.strftime('%Y-%m-%d'))).fetchall()


def proxima_marcacao(regs):
    """RN002 — devolve o próximo tipo permitido na sequência do dia."""
    if not regs:
        return 'entrada'
    ultimo = regs[-1]['tipo']
    if ultimo == 'saida':
        return None
    return SEQUENCIA_PONTO[SEQUENCIA_PONTO.index(ultimo) + 1]


# ---------------------------------------------------------------- Tela 02 dashboard
@app.route('/')
@login_obrigatorio()
def dashboard():
    db = get_db()
    uid, perfil = session['uid'], session['perfil']
    ctx = {}
    regs = marcacoes_do_dia(db, uid)
    ctx['regs_hoje'] = regs
    ctx['proxima'] = proxima_marcacao(regs)
    if perfil == 'colaborador':
        ctx['minhas'] = db.execute(
            'SELECT d.*, g.nome AS criador FROM DBDemanda d '
            'JOIN DBUsuario g ON g.id_usuario=d.id_criador '
            "WHERE d.id_responsavel=? AND d.status<>'concluida' "
            'ORDER BY d.prazo', (uid,)).fetchall()
    else:
        ctx['kpi'] = {
            'colaboradores': db.execute(
                "SELECT COUNT(*) c FROM DBUsuario WHERE perfil='colaborador' AND ativo=1").fetchone()['c'],
            'pendentes': db.execute(
                "SELECT COUNT(*) c FROM DBDemanda WHERE status='pendente'").fetchone()['c'],
            'andamento': db.execute(
                "SELECT COUNT(*) c FROM DBDemanda WHERE status='andamento'").fetchone()['c'],
            'concluidas': db.execute(
                "SELECT COUNT(*) c FROM DBDemanda WHERE status='concluida'").fetchone()['c'],
            'presentes': db.execute(
                "SELECT COUNT(DISTINCT id_usuario) c FROM DBRegistroPonto "
                "WHERE date(data_hora)=date('now','localtime')").fetchone()['c'],
        }
        ctx['urgentes'] = db.execute(
            'SELECT d.*, r.nome AS responsavel FROM DBDemanda d '
            'JOIN DBUsuario r ON r.id_usuario=d.id_responsavel '
            "WHERE d.status<>'concluida' ORDER BY d.prazo LIMIT 6").fetchall()
    return render_template('dashboard.html', **ctx)


# ---------------------------------------------------------------- UC004 Tela 03
@app.route('/ponto', methods=['GET', 'POST'])
@login_obrigatorio()
def ponto():
    db = get_db()
    uid = session['uid']
    regs = marcacoes_do_dia(db, uid)
    proxima = proxima_marcacao(regs)
    if request.method == 'POST':
        tipo = request.form.get('tipo')
        if tipo != proxima:                       # RN002
            flash(MSG['MSG006'], 'danger')
        else:
            inserir_ponto(db, uid, tipo, datetime.now())   # RN003: NSR + hash
            db.commit()
            flash(MSG['MSG001'], 'success')
        return redirect(url_for('ponto'))
    return render_template('registrar_ponto.html', regs=regs, proxima=proxima)


# ---------------------------------------------------------------- UC005 Tela 04
@app.route('/espelho')
@login_obrigatorio()
def espelho():
    db = get_db()
    uid = session['uid']
    mes = request.args.get('mes') or date.today().strftime('%Y-%m')
    regs = db.execute(
        "SELECT * FROM DBRegistroPonto WHERE id_usuario=? AND strftime('%Y-%m',data_hora)=? "
        'ORDER BY data_hora', (uid, mes)).fetchall()
    # agrupa por dia e calcula horas trabalhadas
    dias = {}
    for r in regs:
        d = r['data_hora'][:10]
        dias.setdefault(d, []).append(r)
    espelho_rows, total_min = [], 0
    for d in sorted(dias):
        marc = {m['tipo']: m for m in dias[d]}
        minutos = 0
        try:
            fmt = '%Y-%m-%d %H:%M:%S'
            if 'entrada' in marc and 'pausa' in marc:
                minutos += int((datetime.strptime(marc['pausa']['data_hora'], fmt) -
                                datetime.strptime(marc['entrada']['data_hora'], fmt)).total_seconds() // 60)
            if 'retorno' in marc and 'saida' in marc:
                minutos += int((datetime.strptime(marc['saida']['data_hora'], fmt) -
                                datetime.strptime(marc['retorno']['data_hora'], fmt)).total_seconds() // 60)
        except Exception:
            minutos = 0
        total_min += minutos
        espelho_rows.append({'dia': d, 'marc': marc, 'regs': dias[d], 'minutos': minutos})
    return render_template('espelho_ponto.html', rows=espelho_rows, mes=mes,
                           total_min=total_min)


# ---------------------------------------------------------------- UC008 Tela 05
@app.route('/demandas', methods=['GET', 'POST'])
@login_obrigatorio()
def demandas():
    db = get_db()
    uid = session['uid']
    if request.method == 'POST':
        f = request.form
        if f.get('acao') == 'avancar':                       # RN006 — fluxo sequencial
            d = db.execute('SELECT * FROM DBDemanda WHERE id_demanda=? AND id_responsavel=?',
                           (f['id'], uid)).fetchone()
            if d and d['status'] in PROXIMO_STATUS:
                novo = PROXIMO_STATUS[d['status']]
                db.execute('UPDATE DBDemanda SET status=?, data_conclusao=? WHERE id_demanda=?',
                           (novo, datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            if novo == 'concluida' else None, d['id_demanda']))
                db.commit()
                flash(MSG['MSG002'], 'success')
        elif f.get('id'):                                    # edição da própria demanda
            d = db.execute('SELECT * FROM DBDemanda WHERE id_demanda=? AND id_responsavel=?',
                           (f['id'], uid)).fetchone()
            if d:
                db.execute('UPDATE DBDemanda SET titulo=?, descricao=?, prioridade=?, prazo=? '
                           'WHERE id_demanda=?',
                           (f['titulo'], f['descricao'], f['prioridade'],
                            f['prazo'] or None, f['id']))
                db.commit()
                flash(MSG['MSG002'], 'success')
        else:                                                # criação pelo colaborador (RN005)
            db.execute('INSERT INTO DBDemanda (titulo,descricao,prioridade,status,prazo,'
                       'id_responsavel,id_criador) VALUES (?,?,?,?,?,?,?)',
                       (f['titulo'], f['descricao'], f['prioridade'], 'pendente',
                        f['prazo'] or None, uid, uid))
            db.commit()
            flash(MSG['MSG002'], 'success')
        return redirect(url_for('demandas'))
    filtro = request.args.get('status', 'todas')
    sql = ('SELECT d.*, g.nome AS criador FROM DBDemanda d '
           'JOIN DBUsuario g ON g.id_usuario=d.id_criador WHERE d.id_responsavel=?')
    args = [uid]
    if filtro != 'todas':
        sql += ' AND d.status=?'
        args.append(filtro)
    sql += " ORDER BY CASE d.status WHEN 'pendente' THEN 0 WHEN 'andamento' THEN 1 ELSE 2 END, d.prazo"
    return render_template('acompanhar_demandas.html',
                           demandas=db.execute(sql, args).fetchall(), filtro=filtro)


# ---------------------------------------------------------------- UC007 Tela 06
@app.route('/gerenciar-demandas', methods=['GET', 'POST'])
@login_obrigatorio(('gestor', 'admin'))
def gerenciar_demandas():
    db = get_db()
    if request.method == 'POST':
        f = request.form
        if f.get('id'):
            db.execute('UPDATE DBDemanda SET titulo=?, descricao=?, prioridade=?, '
                       'status=?, prazo=?, id_responsavel=? WHERE id_demanda=?',
                       (f['titulo'], f['descricao'], f['prioridade'], f['status'],
                        f['prazo'] or None, f['id_responsavel'], f['id']))
        else:
            db.execute('INSERT INTO DBDemanda (titulo,descricao,prioridade,status,prazo,'
                       'id_responsavel,id_criador) VALUES (?,?,?,?,?,?,?)',
                       (f['titulo'], f['descricao'], f['prioridade'], 'pendente',
                        f['prazo'] or None, f['id_responsavel'], session['uid']))
        db.commit()
        flash(MSG['MSG002'], 'success')
        return redirect(url_for('gerenciar_demandas'))
    rows = db.execute(
        'SELECT d.*, r.nome AS responsavel FROM DBDemanda d '
        'JOIN DBUsuario r ON r.id_usuario=d.id_responsavel '
        "ORDER BY CASE d.status WHEN 'pendente' THEN 0 WHEN 'andamento' THEN 1 ELSE 2 END, d.prazo").fetchall()
    colabs = db.execute(
        "SELECT id_usuario, nome FROM DBUsuario WHERE perfil='colaborador' AND ativo=1 "
        'ORDER BY nome').fetchall()
    return render_template('gerenciar_demandas.html', demandas=rows, colabs=colabs)


# ---------------------------------------------------------------- UC009 Tela 07
@app.route('/equipe')
@login_obrigatorio(('gestor', 'admin'))
def equipe():
    db = get_db()
    colabs = db.execute(
        "SELECT * FROM DBUsuario WHERE perfil='colaborador' AND ativo=1 ORDER BY nome").fetchall()
    linhas = []
    for c in colabs:
        regs = marcacoes_do_dia(db, c['id_usuario'])
        ultimo = regs[-1] if regs else None
        situacao = ('Sem registro' if not regs else
                    'Jornada encerrada' if ultimo['tipo'] == 'saida' else
                    'Em pausa' if ultimo['tipo'] == 'pausa' else 'Trabalhando')
        dem = db.execute(
            "SELECT SUM(status='pendente') p, SUM(status='andamento') a, "
            "SUM(status='concluida') c FROM DBDemanda WHERE id_responsavel=?",
            (c['id_usuario'],)).fetchone()
        linhas.append({'u': c, 'ultimo': ultimo, 'situacao': situacao,
                       'p': dem['p'] or 0, 'a': dem['a'] or 0, 'c': dem['c'] or 0})
    return render_template('acompanhar_equipe.html', linhas=linhas)




# ---------------------------------------------------------------- UC011 Tela 11
def minutos_trabalhados_mes(db, uid, ano, mes):
    """Soma os minutos trabalhados no mês pareando entrada/pausa/retorno/saída."""
    fmt = '%Y-%m-%d %H:%M:%S'
    regs = db.execute(
        "SELECT * FROM DBRegistroPonto WHERE id_usuario=? AND strftime('%Y-%m',data_hora)=? "
        'ORDER BY data_hora', (uid, f'{ano:04d}-{mes:02d}')).fetchall()
    dias = {}
    for r in regs:
        dias.setdefault(r['data_hora'][:10], {})[r['tipo']] = r
    total = 0
    for marc in dias.values():
        try:
            if 'entrada' in marc and 'pausa' in marc:
                total += int((datetime.strptime(marc['pausa']['data_hora'], fmt) -
                              datetime.strptime(marc['entrada']['data_hora'], fmt)).seconds / 60)
                if 'retorno' in marc and 'saida' in marc:
                    total += int((datetime.strptime(marc['saida']['data_hora'], fmt) -
                                  datetime.strptime(marc['retorno']['data_hora'], fmt)).seconds / 60)
            elif 'entrada' in marc and 'saida' in marc:
                total += int((datetime.strptime(marc['saida']['data_hora'], fmt) -
                              datetime.strptime(marc['entrada']['data_hora'], fmt)).seconds / 60)
        except (ValueError, KeyError):
            continue
    return total


def calcular_aproveitamento(db, uid, ano, mes):
    """RN009 — Índice de aproveitamento do colaborador em home office.

    pct_jornada  = minutos trabalhados / minutos esperados (dias úteis x 8h),
                   considerando apenas os dias já decorridos do mês;
    pct_demandas = demandas concluídas dentro do prazo / (concluídas no mês
                   + demandas em aberto com prazo vencido);
    indice geral = média de pct_jornada e pct_demandas (ou apenas pct_jornada,
                   quando não houver demandas no período avaliado).
    """
    hoje = date.today()
    ultimo_dia = (date(ano + (mes == 12), (mes % 12) + 1, 1) - timedelta(days=1))
    limite = min(hoje, ultimo_dia)
    dias_uteis = sum(1 for i in range(1, limite.day + 1)
                     if date(ano, mes, i).weekday() < 5) if limite.month == mes else 0
    esperado = dias_uteis * 8 * 60
    trabalhado = minutos_trabalhados_mes(db, uid, ano, mes)
    pct_jornada = min(100.0, trabalhado / esperado * 100) if esperado else 0.0

    comp = f'{ano:04d}-{mes:02d}'
    concluidas = db.execute(
        "SELECT prazo, data_conclusao FROM DBDemanda WHERE id_responsavel=? "
        "AND status='concluida' AND strftime('%Y-%m',data_conclusao)=?",
        (uid, comp)).fetchall()
    atrasadas = db.execute(
        "SELECT COUNT(*) c FROM DBDemanda WHERE id_responsavel=? AND status<>'concluida' "
        "AND prazo IS NOT NULL AND prazo < date('now','localtime')", (uid,)).fetchone()['c']
    no_prazo = sum(1 for d in concluidas
                   if d['prazo'] is None or d['data_conclusao'][:10] <= d['prazo'])
    universo = len(concluidas) + atrasadas
    pct_demandas = (no_prazo / universo * 100) if universo else None

    geral = (pct_jornada + pct_demandas) / 2 if pct_demandas is not None else pct_jornada
    return {'pct_jornada': round(pct_jornada, 1),
            'pct_demandas': round(pct_demandas, 1) if pct_demandas is not None else None,
            'pct_geral': round(geral, 1),
            'horas': f'{trabalhado // 60}:{trabalhado % 60:02d}',
            'horas_esperadas': f'{esperado // 60}:{esperado % 60:02d}',
            'no_prazo': no_prazo, 'universo': universo}


@app.route('/aproveitamento')
@login_obrigatorio()
def aproveitamento():
    db = get_db()
    hoje = date.today()
    try:
        ano, mes = map(int, request.args.get('mes', f'{hoje.year}-{hoje.month:02d}').split('-'))
    except ValueError:
        ano, mes = hoje.year, hoje.month

    proprio = dict(calcular_aproveitamento(db, session['uid'], ano, mes),
                   nome=session['nome'])
    equipe_idx = None
    if session['perfil'] in ('gestor', 'admin'):
        membros = db.execute(
            "SELECT id_usuario, nome, email FROM DBUsuario "
            "WHERE perfil='colaborador' AND ativo=1 ORDER BY nome").fetchall()
        equipe_idx = [dict(calcular_aproveitamento(db, m['id_usuario'], ano, mes),
                           nome=m['nome'], email=m['email']) for m in membros]
    return render_template('aproveitamento.html', proprio=proprio,
                           equipe=equipe_idx, mes=f'{ano:04d}-{mes:02d}')


# ---------------------------------------------------------------- UC010 Tela 08
@app.route('/relatorios', methods=['GET', 'POST'])
@login_obrigatorio(('gestor', 'admin'))
def relatorios():
    db = get_db()
    if request.method == 'POST':
        tipo = request.form['tipo']
        ini, fim = request.form['inicio'], request.form['fim']
        try:
            wb = gerar_planilha(db, tipo, ini, fim)
        except ImportError:
            flash('A biblioteca openpyxl não está instalada neste Python. '
                  'Execute no terminal:  python -m pip install openpyxl', 'danger')
            return redirect(url_for('relatorios'))
        except Exception:
            flash(MSG['MSG004'], 'danger')
            return redirect(url_for('relatorios'))
        nome = f'relatorio_{tipo}_{ini}_a_{fim}.xlsx'
        db.execute('INSERT INTO DBRelatorio (tipo,periodo_inicio,periodo_fim,id_gerador,arquivo) '
                   'VALUES (?,?,?,?,?)', (tipo, ini, fim, session['uid'], nome))
        db.commit()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name=nome,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    historico = db.execute(
        'SELECT r.*, u.nome AS gerador FROM DBRelatorio r '
        'JOIN DBUsuario u ON u.id_usuario=r.id_gerador '
        'ORDER BY r.data_geracao DESC LIMIT 10').fetchall()
    hoje = date.today()
    return render_template('gerar_relatorios.html', historico=historico,
                           ini_padrao=(hoje - timedelta(days=30)).strftime('%Y-%m-%d'),
                           fim_padrao=hoje.strftime('%Y-%m-%d'))


def gerar_planilha(db, tipo, ini, fim):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    head_fill = PatternFill('solid', fgColor='0E2A47')
    head_font = Font(color='FFFFFF', bold=True)
    if tipo == 'ponto':
        ws.title = 'Registros de Ponto'
        ws.append(['Colaborador', 'CPF', 'Data/Hora', 'Tipo', 'NSR', 'Hash'])
        rows = db.execute(
            'SELECT u.nome, u.cpf, r.data_hora, r.tipo, r.nsr, r.hash '
            'FROM DBRegistroPonto r JOIN DBUsuario u ON u.id_usuario=r.id_usuario '
            'WHERE date(r.data_hora) BETWEEN ? AND ? ORDER BY u.nome, r.data_hora',
            (ini, fim)).fetchall()
        for r in rows:
            ws.append([r['nome'], r['cpf'], r['data_hora'],
                       ROTULO_PONTO[r['tipo']], r['nsr'], r['hash'][:16] + '…'])
    else:
        ws.title = 'Demandas'
        ws.append(['Título', 'Responsável', 'Prioridade', 'Status', 'Prazo', 'Criada em'])
        rows = db.execute(
            'SELECT d.titulo, u.nome, d.prioridade, d.status, d.prazo, d.data_criacao '
            'FROM DBDemanda d JOIN DBUsuario u ON u.id_usuario=d.id_responsavel '
            'WHERE date(d.data_criacao) <= ? ORDER BY d.prazo', (fim,)).fetchall()
        for r in rows:
            ws.append(list(r))
    for cell in ws[1]:
        cell.fill, cell.font = head_fill, head_font
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            12, min(46, max(len(str(c.value or '')) for c in col) + 2))
    return wb


# ---------------------------------------------------------------- UC002 Tela 09
@app.route('/usuarios', methods=['GET', 'POST'])
@login_obrigatorio(('admin',))
def usuarios():
    db = get_db()
    if request.method == 'POST':
        f = request.form
        if f.get('acao') == 'alternar':
            db.execute('UPDATE DBUsuario SET ativo = 1-ativo WHERE id_usuario=?', (f['id'],))
        elif f.get('id'):
            db.execute('UPDATE DBUsuario SET nome=?, email=?, cpf=?, perfil=? WHERE id_usuario=?',
                       (f['nome'], f['email'].lower(), f['cpf'], f['perfil'], f['id']))
        else:
            db.execute('INSERT INTO DBUsuario (nome,email,cpf,senha,perfil) VALUES (?,?,?,?,?)',
                       (f['nome'], f['email'].lower(), f['cpf'],
                        generate_password_hash(f.get('senha') or '123456'), f['perfil']))
        db.commit()
        flash(MSG['MSG003'], 'success')
        return redirect(url_for('usuarios'))
    rows = db.execute('SELECT * FROM DBUsuario ORDER BY nome').fetchall()
    return render_template('gerenciar_usuarios.html', usuarios=rows)


# ---------------------------------------------------------------- UC006 Tela 10
@app.route('/exportar-afd', methods=['GET', 'POST'])
@login_obrigatorio(('admin',))
def exportar_afd():
    db = get_db()
    hoje = date.today()
    if request.method == 'POST':
        ini, fim = request.form['inicio'], request.form['fim']
        try:
            conteudo = gerar_afd(db, ini, fim)
        except Exception:
            flash(MSG['MSG004'], 'danger')
            return redirect(url_for('exportar_afd'))
        buf = io.BytesIO(conteudo.encode('utf-8'))
        return send_file(buf, as_attachment=True,
                         download_name=f'AFD_GlobalControle_{ini}_a_{fim}.txt',
                         mimetype='text/plain')
    total = db.execute('SELECT COUNT(*) c FROM DBRegistroPonto').fetchone()['c']
    ultimo = db.execute('SELECT MAX(data_hora) m FROM DBRegistroPonto').fetchone()['m']
    return render_template('exportar_afd.html', total=total, ultimo=ultimo,
                           ini_padrao=hoje.replace(day=1).strftime('%Y-%m-%d'),
                           fim_padrao=hoje.strftime('%Y-%m-%d'))


def gerar_afd(db, ini, fim):
    """Leiaute simplificado do AFD para REP-P (Portaria nº 671/2021):
    registro tipo 1 (cabeçalho), tipo 7 (marcações REP-P) e trailer 9."""
    linhas = []
    nsr_global = 1
    cab = (f'{0:09d}1' + '1' + '12345678000199'.ljust(14) +
           'GLOBALWEB CORP'.ljust(150)[:150] +
           ini.replace('-', '') + fim.replace('-', '') +
           datetime.now().strftime('%d%m%Y%H%M'))
    linhas.append(cab)
    regs = db.execute(
        'SELECT r.*, u.cpf FROM DBRegistroPonto r '
        'JOIN DBUsuario u ON u.id_usuario=r.id_usuario '
        'WHERE date(r.data_hora) BETWEEN ? AND ? ORDER BY r.data_hora', (ini, fim)).fetchall()
    for r in regs:
        dt = datetime.strptime(r['data_hora'], '%Y-%m-%d %H:%M:%S')
        cpf = r['cpf'].replace('.', '').replace('-', '')
        linhas.append(f'{nsr_global:09d}7' + dt.strftime('%d%m%Y%H%M') +
                      cpf.rjust(12, '0') + r['hash'][:64])
        nsr_global += 1
    linhas.append(f'{nsr_global:09d}9' + f'{len(regs):09d}')
    return '\r\n'.join(linhas) + '\r\n'


# ---------------------------------------------------------------- UC003 perfil
@app.route('/perfil', methods=['GET', 'POST'])
@login_obrigatorio()
def perfil():
    db = get_db()
    u = db.execute('SELECT * FROM DBUsuario WHERE id_usuario=?', (session['uid'],)).fetchone()
    if request.method == 'POST':
        f = request.form
        if f.get('senha'):
            db.execute('UPDATE DBUsuario SET nome=?, senha=? WHERE id_usuario=?',
                       (f['nome'], generate_password_hash(f['senha']), session['uid']))
        else:
            db.execute('UPDATE DBUsuario SET nome=? WHERE id_usuario=?',
                       (f['nome'], session['uid']))
        db.commit()
        session['nome'] = f['nome']
        flash(MSG['MSG003'], 'success')
        return redirect(url_for('perfil'))
    return render_template('perfil.html', u=u)


if __name__ == '__main__':
    seed()
    app.run(debug=True, host='0.0.0.0', port=5000)
